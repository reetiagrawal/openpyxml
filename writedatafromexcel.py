#WRITE DATA TO THE EXCEL
'''import openpyxl

wb = openpyxl.load_workbook("/home/shiv/Downloads/openxml.xlsx")
sheet_obj = wb['Name']
row = sheet_obj.max_row
column = sheet_obj.max_column

for i in range(1,row+1):
    for j in range(1,column+1):
        print(sheet_obj.cell(i,j).value)
sheet_obj.cell(row=7,column=1,value="Richi")
sheet_obj.cell(row=7,column=2,value="BE")
sheet_obj.cell(row=7,column=3,value="ECE")
sheet_obj.cell(row=7,column=4,value="8")
wb.save("Report.xlsx")
'''

from openpyxl import Workbook
from openpyxl.styles import PatternFill
wb = Workbook()
wb['Sheet'].title = "Report of Automation"
sheet_obj = wb.active
sheet_obj['A1'].value= "Name"
sheet_obj['B1'].value = "Status"
sheet_obj['A2'].value = "Python"
sheet_obj['B2'].value = "Active"
sheet_obj['B2'].fill = PatternFill("solid",fgColor="71FF33")
sheet_obj['A3'].value = "JAVA"
sheet_obj['B3'].value = "Inactive"
sheet_obj['B3'].fill = PatternFill("solid",fgColor="F50707")
wb.save("/home/shiv/Downloads/openxml1.xlsx")



